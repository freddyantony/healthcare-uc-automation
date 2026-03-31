"""
Multi-Cluster CUCM Gateway IP Report via AXL (v6)
Pulls IP addresses for H.323, MGCP, SIP Trunk, and SCCP Analog VG gateways.
Groups output by device type across multiple clusters into a single workbook.

Requirements:
    pip install requests lxml openpyxl urllib3
"""

import re
import warnings
from collections import defaultdict
from datetime import datetime
import xml.etree.ElementTree as ET

import urllib3
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ──────────────────────────────────────────────────
AXL_VERSION = "14.0" # Match your CUCM version
OUTPUT_FILE = "Gateway_Sip_Trunk_IP_Discovery_Report.xlsx"

CLUSTER_GROUPS = [
    {
        "user": "api_admin",        # <-- UPDATE THIS
        "pass": "your_password",    # <-- UPDATE THIS
        "clusters": [
            {"name": "Cluster-US-East", "ip": "cucm-pub-01.example.local"},
            {"name": "Cluster-US-West", "ip": "cucm-pub-02.example.local"},
        ]
    }
]
# ───────────────────────────────────────────────────────────────────

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

SOAP_ENVELOPE = """<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:ns="http://www.cisco.com/AXL/API/{version}">
    <soapenv:Header/>
    <soapenv:Body>
        {body}
    </soapenv:Body>
</soapenv:Envelope>"""

def axl_sql_query(sql, host, user, pwd, label=""):
    axl_url = f"https://{host}:8443/axl/"
    headers = {
        "Content-Type": "text/xml; charset=utf-8",
        "SOAPAction": f"CUCM:DB ver={AXL_VERSION}",
    }
    body = f"<ns:executeSQLQuery><sql>{sql}</sql></ns:executeSQLQuery>"
    payload = SOAP_ENVELOPE.format(version=AXL_VERSION, body=body)
    
    try:
        resp = requests.post(
            axl_url, data=payload, headers=headers,
            auth=(user, pwd), verify=False, timeout=60,
        )
        resp.raise_for_status()
        tree = etree.fromstring(resp.content)
        rows = []
        for elem in tree.iter():
            local = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if local == "row":
                row = {}
                for child in elem:
                    ctag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                    row[ctag] = child.text.strip() if child.text else ""
                rows.append(row)
        return rows
    except Exception as e:
        print(f"    [{label}] Error: {e}")
        return []

def risport_get_ips(device_names, host, user, pwd, status="Any"):
    ris_url = f"https://{host}:8443/realtimeservice2/services/RISService70"
    result = {}
    batch_size = 200

    for i in range(0, len(device_names), batch_size):
        batch = device_names[i:i + batch_size]
        items_xml = ""
        for name in batch:
            items_xml += f"<soap:item><soap:Item>{name}</soap:Item></soap:item>\n"

        payload = f"""<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:soap="http://schemas.cisco.com/ast/soap">
   <soapenv:Header/>
   <soapenv:Body>
      <soap:selectCmDevice>
         <soap:StateInfo></soap:StateInfo>
         <soap:CmSelectionCriteria>
            <soap:MaxReturnedDevices>1000</soap:MaxReturnedDevices>
            <soap:DeviceClass>Any</soap:DeviceClass>
            <soap:Model>255</soap:Model>
            <soap:Status>{status}</soap:Status>
            <soap:NodeName></soap:NodeName>
            <soap:SelectBy>Name</soap:SelectBy>
            <soap:SelectItems>
{items_xml}            </soap:SelectItems>
            <soap:Protocol>Any</soap:Protocol>
            <soap:DownloadStatus>Any</soap:DownloadStatus>
         </soap:CmSelectionCriteria>
      </soap:selectCmDevice>
   </soapenv:Body>
</soapenv:Envelope>"""

        try:
            resp = requests.post(
                ris_url, data=payload,
                headers={"Content-Type": "text/xml", "SOAPAction": "selectCmDevice"},
                auth=(user, pwd), verify=False, timeout=60,
            )
            resp.raise_for_status()
            root = ET.fromstring(resp.text)
            ns = {"ns": "http://schemas.cisco.com/ast/soap"}

            for device in root.findall(".//ns:CmDevices/ns:item", ns):
                name_elem = device.find("ns:Name", ns)
                ip_elem = device.find(".//ns:IPAddress/ns:item/ns:IP", ns)
                name = name_elem.text.strip() if name_elem is not None and name_elem.text else ""
                ip = ip_elem.text.strip() if ip_elem is not None and ip_elem.text else ""
                if name and ip and ip != "0.0.0.0":
                    result[name] = ip
        except Exception as e:
            print(f"    [WARN] RisPort batch {i//batch_size + 1} failed: {e}")

    return result

def axl_get_sip_trunk_destinations(trunk_name, host, user, pwd):
    axl_url = f"https://{host}:8443/axl/"
    headers = {
        "Content-Type": "text/xml; charset=utf-8",
        "SOAPAction": f"CUCM:DB ver={AXL_VERSION}",
    }
    body = f"""<ns:getSipTrunk>
        <name>{trunk_name}</name>
        <returnedTags>
            <destinations>
                <destination>
                    <addressIpv4/>
                    <port/>
                    <sortOrder/>
                </destination>
            </destinations>
            <sipProfileName/>
            <securityProfileName/>
        </returnedTags>
    </ns:getSipTrunk>"""
    payload = SOAP_ENVELOPE.format(version=AXL_VERSION, body=body)
    try:
        resp = requests.post(
            axl_url, data=payload, headers=headers,
            auth=(user, pwd), verify=False, timeout=30,
        )
        resp.raise_for_status()
        tree = etree.fromstring(resp.content)

        destinations = []
        sip_profile = ""
        sec_profile = ""
        for elem in tree.iter():
            local = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if local == "addressIpv4" and elem.text:
                destinations.append(elem.text.strip())
            if local == "sipProfileName" and elem.text:
                sip_profile = elem.text.strip()
            if local == "securityProfileName" and elem.text:
                sec_profile = elem.text.strip()

        return {
            "destinations": ", ".join(destinations) if destinations else "N/A",
            "sip_profile": sip_profile,
            "sec_profile": sec_profile,
        }
    except Exception as e:
        return {"destinations": f"Error: {e}", "sip_profile": "", "sec_profile": ""}

def get_h323_gateways(cluster_name, host, user, pwd):
    sql = """SELECT d.name as device_name, d.description,
             dp.name as device_pool, css.name as css_name,
             tp.name as product_type
             FROM device d
             INNER JOIN typeproduct tp ON d.tkmodel = tp.tkmodel
             INNER JOIN devicepool dp ON d.fkdevicepool = dp.pkid
             LEFT JOIN callingsearchspace css ON d.fkcallingsearchspace_restrict = css.pkid
             WHERE tp.name LIKE '%H.323%'"""
    rows = axl_sql_query(sql, host, user, pwd, "H323")
    results = []
    for row in rows:
        name = row.get("device_name", "")
        results.append({
            "Cluster Name": cluster_name,
            "Gateway Name": name,
            "IP Address": name,
            "Description": row.get("description", ""),
            "Device Pool": row.get("device_pool", ""),
            "CSS": row.get("css_name", ""),
            "Product Type": row.get("product_type", ""),
        })
    return results

def get_mgcp_gateways(cluster_name, host, user, pwd):
    sql = """SELECT DISTINCT d.name as device_name, d.description,
             dp.name as device_pool, tp.name as product_type
             FROM device d
             INNER JOIN typeproduct tp ON d.tkmodel = tp.tkmodel
             INNER JOIN devicepool dp ON d.fkdevicepool = dp.pkid
             WHERE d.tkclass = 2
             AND tp.name NOT LIKE '%H.323%'
             ORDER BY d.name"""
    rows = axl_sql_query(sql, host, user, pwd, "MGCP")
    results = []
    for row in rows:
        name = row.get("device_name", "")
        results.append({
            "Cluster Name": cluster_name,
            "Gateway Name": name,
            "IP Address": name,
            "Description": row.get("description", ""),
            "Device Pool": row.get("device_pool", ""),
            "Product Type": row.get("product_type", ""),
        })
    return results

def get_sip_trunks(cluster_name, host, user, pwd):
    sql = """SELECT d.name as trunk_name, d.description,
             dp.name as device_pool, css.name as css_name,
             tp.name as product_type
             FROM device d
             INNER JOIN typeproduct tp ON d.tkmodel = tp.tkmodel
             INNER JOIN devicepool dp ON d.fkdevicepool = dp.pkid
             LEFT JOIN callingsearchspace css ON d.fkcallingsearchspace_restrict = css.pkid
             WHERE tp.name = 'SIP Trunk'
             ORDER BY d.name"""
    rows = axl_sql_query(sql, host, user, pwd, "SIP_trunks")

    if not rows:
        sql2 = """SELECT d.name as trunk_name, d.description,
                  dp.name as device_pool, css.name as css_name,
                  tp.name as product_type
                  FROM device d
                  INNER JOIN typeproduct tp ON d.tkmodel = tp.tkmodel
                  INNER JOIN devicepool dp ON d.fkdevicepool = dp.pkid
                  LEFT JOIN callingsearchspace css ON d.fkcallingsearchspace_restrict = css.pkid
                  WHERE tp.name LIKE '%Trunk%'
                  AND tp.name NOT LIKE '%MTP%'
                  AND tp.name NOT LIKE '%Conference%'
                  AND tp.name NOT LIKE '%Music%'
                  ORDER BY d.name"""
        rows = axl_sql_query(sql2, host, user, pwd, "SIP_trunks_broad")

    results = []
    for idx, row in enumerate(rows):
        trunk_name = row.get("trunk_name", "")
        detail = axl_get_sip_trunk_destinations(trunk_name, host, user, pwd)
        results.append({
            "Cluster Name": cluster_name,
            "Trunk Name": trunk_name,
            "Destination Address(es)": detail["destinations"],
            "Description": row.get("description", ""),
            "Device Pool": row.get("device_pool", ""),
            "CSS": row.get("css_name", ""),
            "SIP Profile": detail["sip_profile"],
            "Security Profile": detail["sec_profile"],
        })
    return results

def decode_an_mac_portion(device_name):
    raw = device_name.upper().replace("AN", "", 1)
    if len(raw) >= 3:
        mac_portion = raw[:-3]
        mac_padded = mac_portion.ljust(12, "0")[:12]
        mac_formatted = ":".join(mac_padded[i:i+2] for i in range(0, 12, 2))
        return mac_portion, mac_formatted
    return raw, raw

def get_analog_vg_ips(cluster_name, host, user, pwd):
    sql = """SELECT d.name as device_name, d.description, dp.name as device_pool
             FROM device d
             INNER JOIN devicepool dp ON d.fkdevicepool = dp.pkid
             WHERE d.name LIKE 'AN%'
             AND d.tkclass = 1
             ORDER BY d.description"""
    rows = axl_sql_query(sql, host, user, pwd, "AN_devices")

    vg_groups = defaultdict(list)
    for row in rows:
        name = row.get("device_name", "")
        mac_portion, mac_formatted = decode_an_mac_portion(name)
        vg_groups[mac_formatted].append({
            "device_name": name,
            "description": row.get("description", ""),
            "device_pool": row.get("device_pool", ""),
            "mac_portion": mac_portion,
        })

    sample_to_mac = {devices[0]["device_name"]: mac for mac, devices in vg_groups.items()}
    all_samples = list(sample_to_mac.keys())
    
    ris_ips = risport_get_ips(all_samples, host, user, pwd, status="Any") if all_samples else {}
    
    vg_ip_map = {}
    for sample_name, ip in ris_ips.items():
        mac = sample_to_mac.get(sample_name)
        if mac:
            vg_ip_map[mac] = ip

    results = []
    for mac, devices in sorted(vg_groups.items()):
        desc = devices[0].get("description", "")
        vg_name = desc
        port_match = re.search(r'\s+\d+/\d+/\d+\s*$', desc)
        if port_match:
            vg_name = desc[:port_match.start()].strip()
        elif " " in desc:
            vg_name = desc.rsplit(" ", 1)[0].strip()
        if not vg_name:
            vg_name = mac

        ip = vg_ip_map.get(mac, "")
        results.append({
            "Cluster Name": cluster_name,
            "VG Name": vg_name,
            "IP Address": ip,
            "VG MAC": mac,
            "Sample AN Device": devices[0]["device_name"],
            "Device Pool": devices[0]["device_pool"],
            "Port Count": len(devices),
            "Note": "Resolved via RIS" if ip else "Not registered / no siblings online"
        })
    return results

def write_excel(h323_data, mgcp_data, sip_data, vg_data):
    wb = Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    af = PatternFill("solid", fgColor="D6E4F0")

    def write_sheet(ws, title, headers, data):
        ws.title = title
        ws.sheet_properties.tabColor = "2F5496"
        ws.freeze_panes = "A2"
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb
        for ri, rd in enumerate(data, 2):
            for ci, h in enumerate(headers, 1):
                v = rd.get(h, "")
                c = ws.cell(row=ri, column=ci, value=v)
                c.font, c.alignment, c.border = cf, ca, tb
                if ri % 2 == 0:
                    c.fill = af
        for ci in range(1, len(headers) + 1):
            ml = len(str(headers[ci - 1]))
            for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 2),
                                    min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value:
                        ml = max(ml, min(len(str(cell.value)), 60))
            ws.column_dimensions[get_column_letter(ci)].width = ml + 4
        if not data:
            ws.cell(row=2, column=1, value="No data found").font = Font(
                name="Arial", italic=True, color="888888")

    write_sheet(wb.active, "H.323 Gateways",
                ["Cluster Name", "Gateway Name", "IP Address", "Description",
                 "Device Pool", "CSS", "Product Type"], h323_data)
    write_sheet(wb.create_sheet(), "MGCP Gateways",
                ["Cluster Name", "Gateway Name", "IP Address", "Description",
                 "Device Pool", "Product Type"], mgcp_data)
    write_sheet(wb.create_sheet(), "SIP Trunks",
                ["Cluster Name", "Trunk Name", "Destination Address(es)", "Description",
                 "Device Pool", "CSS", "SIP Profile", "Security Profile"],
                sip_data)
    write_sheet(wb.create_sheet(), "Analog VG (SCCP)",
                ["Cluster Name", "VG Name", "IP Address", "VG MAC", "Sample AN Device",
                 "Device Pool", "Port Count", "Note"], vg_data)

    sws = wb.create_sheet("Summary", 0)
    sws.sheet_properties.tabColor = "00B050"
    sd = [
        {"Gateway Type": "H.323 Gateways", "Total Count": len(h323_data)},
        {"Gateway Type": "MGCP Gateways", "Total Count": len(mgcp_data)},
        {"Gateway Type": "SIP Trunks", "Total Count": len(sip_data)},
        {"Gateway Type": "Analog VG (SCCP)", "Total Count": len(vg_data)},
    ]
    write_sheet(sws, "Summary", ["Gateway Type", "Total Count"], sd)
    sws.sheet_properties.tabColor = "00B050"
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    sws.cell(row=7, column=1, value=f"Generated: {ts}").font = Font(
        name="Arial", italic=True, size=9, color="666666")
    wb.save(OUTPUT_FILE)
    print(f"\n[+] Global Report saved: {OUTPUT_FILE}")

def main():
    print("=" * 60)
    print("  Multi-Cluster Gateway & Trunk Discovery")
    print("=" * 60)

    master_h323 = []
    master_mgcp = []
    master_sip = []
    master_vg = []

    for group in CLUSTER_GROUPS:
        axl_user = group["user"]
        axl_pass = group["pass"]
        
        for cluster in group["clusters"]:
            cluster_name = cluster["name"]
            cucm_ip = cluster["ip"]
            
            print(f"\n[{cluster_name}] Connecting to {cucm_ip}...")
            
            test = axl_sql_query("SELECT count(*) as total FROM device", cucm_ip, axl_user, axl_pass, "TEST")
            if not test:
                print(f"    ❌ FAILED - Skipping {cluster_name}. Check IP/Credentials.")
                continue

            print(f"    ✅ Connected. Total devices in DB: {test[0].get('total', '?')}")
            
            master_h323.extend(get_h323_gateways(cluster_name, cucm_ip, axl_user, axl_pass))
            master_mgcp.extend(get_mgcp_gateways(cluster_name, cucm_ip, axl_user, axl_pass))
            master_sip.extend(get_sip_trunks(cluster_name, cucm_ip, axl_user, axl_pass))
            master_vg.extend(get_analog_vg_ips(cluster_name, cucm_ip, axl_user, axl_pass))

    write_excel(master_h323, master_mgcp, master_sip, master_vg)

if __name__ == "__main__":
    main()